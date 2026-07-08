from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
import io
import json
import logging
import unicodedata
from functools import wraps
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.pdfgen.canvas import Canvas
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from dotenv import load_dotenv

load_dotenv()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY')
if not app.secret_key:
    app.secret_key = 'comparador_inventario_2024_secret'
    logger.warning(
        'SECRET_KEY não definida no ambiente — usando chave padrão insegura. '
        'Defina SECRET_KEY (ex.: em um arquivo .env) antes de ir para produção.'
    )

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def extensao_permitida(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

UNIDADES = [
    'Pesqueira', 'Caruaru', 'Garanhuns Centro', 'Garanhuns BR',
    'Santo Antônio', 'Cruza', 'Maceió', 'CPG', 'Filial de CPG', 'Rancho CD', 'Natal'
]
CONFIGURACOES_PADRAO = {
    'limiar_alerta_critico': '80',
    'limiar_alerta_aviso': '95',
    'limite_historico': '20',
    'limite_resultados': '1000'
}

def init_db():
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()

    # Tabela de usuários
    c.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        senha TEXT NOT NULL,
        nome TEXT NOT NULL,
        email TEXT,
        unidade TEXT NOT NULL,
        é_admin BOOLEAN DEFAULT 0,
        ativo BOOLEAN DEFAULT 1,
        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # Tabela de comparações
    c.execute('''CREATE TABLE IF NOT EXISTS comparacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER NOT NULL,
        unidade TEXT NOT NULL,
        data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        arquivo_inventario TEXT,
        arquivo_posicao TEXT,
        total_itens INTEGER,
        itens_iguais INTEGER,
        itens_diferentes INTEGER,
        itens_falta INTEGER,
        resultado_json TEXT,
        FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
    )''')

    # Tabela de detalhes das comparações
    c.execute('''CREATE TABLE IF NOT EXISTS comparacao_detalhes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        comparacao_id INTEGER NOT NULL,
        codigo TEXT,
        descricao TEXT,
        loja TEXT,
        qtd_inv REAL,
        valor_inv REAL,
        qtd_pos REAL,
        valor_pos REAL,
        diff_qtd REAL,
        diff_valor REAL,
        status TEXT,
        FOREIGN KEY(comparacao_id) REFERENCES comparacoes(id)
    )''')

    # Tabela de alertas
    c.execute('''CREATE TABLE IF NOT EXISTS alertas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        comparacao_id INTEGER NOT NULL,
        usuario_id INTEGER NOT NULL,
        tipo_alerta TEXT,
        mensagem TEXT,
        taxa_acuracidade REAL,
        limite_alerta REAL,
        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        lido BOOLEAN DEFAULT 0,
        FOREIGN KEY(comparacao_id) REFERENCES comparacoes(id),
        FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS configuracoes (
        chave TEXT PRIMARY KEY,
        valor TEXT NOT NULL,
        atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS atividades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario_id INTEGER,
        tipo TEXT NOT NULL,
        mensagem TEXT NOT NULL,
        criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
    )''')

    for chave, valor in CONFIGURACOES_PADRAO.items():
        c.execute(
            'INSERT OR IGNORE INTO configuracoes (chave, valor) VALUES (?, ?)',
            (chave, valor)
        )

    conn.commit()

    # Criar usuário demo se não existir
    c.execute('SELECT * FROM usuarios WHERE username = ?', ('admin',))
    if not c.fetchone():
        senha_hash = generate_password_hash('admin123')
        c.execute('''INSERT INTO usuarios (username, senha, nome, unidade, é_admin)
                     VALUES (?, ?, ?, ?, ?)''',
                  ('admin', senha_hash, 'Administrador', 'Caruaru', 1))
        conn.commit()

    conn.close()

init_db()

def obter_configuracao(chave, conversor=str):
    padrao = CONFIGURACOES_PADRAO[chave]
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('SELECT valor FROM configuracoes WHERE chave = ?', (chave,))
    row = c.fetchone()
    conn.close()

    try:
        return conversor(row[0] if row else padrao)
    except (TypeError, ValueError):
        return conversor(padrao)

def registrar_atividade(usuario_id, tipo, mensagem):
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute(
        'INSERT INTO atividades (usuario_id, tipo, mensagem) VALUES (?, ?, ?)',
        (usuario_id, tipo, mensagem)
    )
    conn.commit()
    conn.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def verificar_admin():
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('SELECT é_admin FROM usuarios WHERE id = ?', (session.get('usuario_id'),))
    user = c.fetchone()
    conn.close()
    return bool(user and user[0])

def normalizar_rotulo(valor):
    if pd.isna(valor):
        return ''
    texto = unicodedata.normalize('NFKD', str(valor))
    return ''.join(char for char in texto if not unicodedata.combining(char)).strip().casefold()

def encontrar_layout_posicao_resumida(df):
    """Localiza as colunas do relatório resumido de posição diária."""
    limite = min(len(df), 30)

    for indice in range(limite):
        colunas = {}
        for numero_coluna, valor in enumerate(df.iloc[indice]):
            rotulo = normalizar_rotulo(valor)
            if not rotulo:
                continue

            if 'codigo' in rotulo and 'codigo' not in colunas:
                colunas['codigo'] = numero_coluna
            elif 'descricao' in rotulo and 'descricao' not in colunas:
                colunas['descricao'] = numero_coluna
            elif (rotulo.startswith('qtde') or 'quantidade' in rotulo) and 'quantidade' not in colunas:
                colunas['quantidade'] = numero_coluna
            elif 'custo medio' in rotulo and 'custo_medio' not in colunas:
                colunas['custo_medio'] = numero_coluna
            elif (rotulo == 'total' or 'valor total' in rotulo) and 'valor_total' not in colunas:
                colunas['valor_total'] = numero_coluna

        if {'codigo', 'descricao', 'quantidade', 'custo_medio', 'valor_total'} <= colunas.keys():
            return indice, colunas

    return None

def ler_aba_compativel(file_path, tipo, validador):
    """Lê a aba compatível sem depender do nome usado no Excel."""
    try:
        arquivo_excel = pd.ExcelFile(file_path)
    except Exception as erro:
        raise ValueError(f'Não foi possível abrir o arquivo de {tipo}: {erro}') from erro

    try:
        nomes_abas = arquivo_excel.sheet_names
        if not nomes_abas:
            raise ValueError(f'O arquivo de {tipo} não possui planilhas')

        # Mantém compatibilidade com o modelo antigo, mas tenta todas as abas.
        abas_ordenadas = sorted(nomes_abas, key=lambda nome: nome != 'Planilha1')
        for nome_aba in abas_ordenadas:
            df = pd.read_excel(arquivo_excel, sheet_name=nome_aba, header=None)
            if validador(df):
                logger.info(f'Aba selecionada para {tipo}: {nome_aba}')
                return df

        abas = ', '.join(nomes_abas)
        raise ValueError(
            f'Nenhuma aba compatível foi encontrada no arquivo de {tipo}. '
            f'Abas encontradas: {abas}'
        )
    finally:
        arquivo_excel.close()

def validar_coluna_extraida(serie, campo, tipo, contexto):
    """Verifica se uma coluna extraída realmente contém o tipo de dado esperado.

    Esses relatórios usam células mescladas no cabeçalho que deslocam a coluna
    de dados de forma inconsistente entre arquivos (às vezes 1 coluna para a
    esquerda, às vezes para a direita). Um índice de coluna correto para um
    arquivo pode silenciosamente apontar para a coluna errada em outro — o
    sintoma nesse caso é a comparação inteira sair como "DIFERENCA"/vazia sem
    nenhum erro visível. Esta checagem transforma isso em um erro imediato e
    específico, em vez de um resultado silenciosamente errado.
    """
    valores = serie.dropna()
    if valores.empty:
        raise ValueError(
            f'A coluna "{campo}" do arquivo de {contexto} veio completamente vazia. '
            'O layout do arquivo provavelmente mudou — confira as posições de coluna '
            'usadas na extração.'
        )

    if tipo == 'numerico':
        proporcao_valida = pd.to_numeric(valores, errors='coerce').notna().mean()
    else:
        proporcao_valida = valores.astype(str).str.contains(r'[A-Za-zÀ-ÿ]', regex=True).mean()

    if proporcao_valida < 0.5:
        tipo_esperado = 'números' if tipo == 'numerico' else 'texto'
        raise ValueError(
            f'A coluna "{campo}" do arquivo de {contexto} não parece conter {tipo_esperado} '
            f'válidos (apenas {proporcao_valida:.0%} das linhas batem). O layout do arquivo '
            'provavelmente mudou — confira as posições de coluna usadas na extração.'
        )

def extrair_inventario(file_path):
    df = ler_aba_compativel(
        file_path,
        'inventário',
        lambda planilha: planilha.shape[0] > 6 and planilha.shape[1] >= 14
    )
    dados = df.iloc[6:].copy()

    dados_limpos = pd.DataFrame({
        'codigo': dados.iloc[:, 2],
        'loja': dados.iloc[:, 4],
        'descricao': dados.iloc[:, 5],
        'unidade': dados.iloc[:, 8],
        'quantidade': dados.iloc[:, 10],
        'valor_unitario': dados.iloc[:, 11],
        'valor_total': dados.iloc[:, 13]
    })

    dados_limpos = dados_limpos.dropna(how='all')
    dados_limpos['codigo'] = pd.to_numeric(dados_limpos['codigo'], errors='coerce')
    dados_limpos['loja'] = pd.to_numeric(dados_limpos['loja'], errors='coerce')
    dados_limpos['quantidade'] = pd.to_numeric(dados_limpos['quantidade'], errors='coerce')
    dados_limpos['valor_unitario'] = pd.to_numeric(dados_limpos['valor_unitario'], errors='coerce')
    dados_limpos['valor_total'] = pd.to_numeric(dados_limpos['valor_total'], errors='coerce')

    dados_limpos = dados_limpos.dropna(subset=['codigo'])
    dados_limpos = dados_limpos[(dados_limpos['codigo'] > 0) & (dados_limpos['codigo'] <= 99999)]

    validar_coluna_extraida(dados_limpos['descricao'], 'Descrição Produto', 'texto', 'inventário')
    validar_coluna_extraida(dados_limpos['quantidade'], 'Quantidade', 'numerico', 'inventário')
    validar_coluna_extraida(dados_limpos['valor_total'], 'Valor Parcial', 'numerico', 'inventário')

    dados_limpos['codigo_formatado'] = dados_limpos['codigo'].astype(int).astype(str).str.zfill(7)
    dados_limpos['chave'] = dados_limpos['codigo_formatado'] + '_' + dados_limpos['loja'].astype(str)

    if dados_limpos.empty:
        raise ValueError('O arquivo de inventário não contém produtos válidos')

    return dados_limpos

def extrair_posicao(file_path, loja_padrao=None):
    df = ler_aba_compativel(
        file_path,
        'posição diária',
        lambda planilha: (
            planilha.shape[1] >= 12
            or encontrar_layout_posicao_resumida(planilha) is not None
        )
    )

    if df.shape[1] >= 12:
        dados = df.iloc[4:].copy()

        # O cabeçalho (linha 3) às vezes tem uma coluna em branco extra antes do "Total"
        # (célula mesclada), o que desloca a coluna de valor. Localiza pelo rótulo em vez
        # de assumir sempre a coluna 11.
        coluna_valor_total = 11
        for numero_coluna, valor in enumerate(df.iloc[3]):
            rotulo = normalizar_rotulo(valor)
            if rotulo == 'total' or 'valor total' in rotulo:
                coluna_valor_total = numero_coluna
                break

        dados_limpos = pd.DataFrame({
            'codigo': dados.iloc[:, 0],
            'descricao': dados.iloc[:, 2],
            'unidade': dados.iloc[:, 6],
            'loja': dados.iloc[:, 7],
            'custo_medio': dados.iloc[:, 9],
            'quantidade': dados.iloc[:, 10],
            'valor_total': dados.iloc[:, coluna_valor_total]
        })
    else:
        layout = encontrar_layout_posicao_resumida(df)
        if not layout:
            raise ValueError('O formato resumido da posição diária não foi reconhecido')
        if loja_padrao is None:
            raise ValueError(
                'A posição diária não informa a loja e o inventário possui mais de uma loja'
            )

        linha_cabecalho, colunas = layout
        dados = df.iloc[linha_cabecalho + 1:].copy()
        dados_limpos = pd.DataFrame({
            'codigo': dados.iloc[:, colunas['codigo']],
            'descricao': dados.iloc[:, colunas['descricao']],
            'unidade': '',
            'loja': loja_padrao,
            'custo_medio': dados.iloc[:, colunas['custo_medio']],
            'quantidade': dados.iloc[:, colunas['quantidade']],
            'valor_total': dados.iloc[:, colunas['valor_total']]
        })

    dados_limpos = dados_limpos.dropna(how='all')
    dados_limpos['codigo'] = pd.to_numeric(dados_limpos['codigo'], errors='coerce')
    dados_limpos['loja'] = pd.to_numeric(dados_limpos['loja'], errors='coerce')
    dados_limpos['quantidade'] = pd.to_numeric(dados_limpos['quantidade'], errors='coerce')
    dados_limpos['custo_medio'] = pd.to_numeric(dados_limpos['custo_medio'], errors='coerce')
    dados_limpos['valor_total'] = pd.to_numeric(dados_limpos['valor_total'], errors='coerce')

    dados_limpos = dados_limpos.dropna(subset=['codigo'])
    dados_limpos = dados_limpos[(dados_limpos['codigo'] > 0) & (dados_limpos['codigo'] <= 99999)]

    validar_coluna_extraida(dados_limpos['descricao'], 'Descrição', 'texto', 'posição diária')
    validar_coluna_extraida(dados_limpos['quantidade'], 'Quantidade', 'numerico', 'posição diária')
    validar_coluna_extraida(dados_limpos['valor_total'], 'Total', 'numerico', 'posição diária')

    dados_limpos['codigo_formatado'] = dados_limpos['codigo'].astype(int).astype(str).str.zfill(7)
    dados_limpos['chave'] = dados_limpos['codigo_formatado'] + '_' + dados_limpos['loja'].astype(str)

    if dados_limpos.empty:
        raise ValueError('O arquivo de posição diária não contém produtos válidos')

    return dados_limpos

def comparar(inv_data, pos_data):
    comparacao = pd.merge(
        inv_data,
        pos_data,
        on='chave',
        how='outer',
        suffixes=('_inv', '_pos')
    )

    # Quando o item existe só de um lado (FALTA NO INVENTARIO/NA POSICAO), o
    # relatório usa sempre as colunas "_inv" para identificar o produto. Sem
    # isso, um item que falta no inventário aparece sem código/descrição/loja
    # nenhuma, mesmo essas informações existindo do lado da posição.
    for campo in ('codigo_formatado', 'descricao', 'loja'):
        comparacao[f'{campo}_inv'] = comparacao[f'{campo}_inv'].fillna(comparacao[f'{campo}_pos'])

    # A comparação segue a mesma precisão exibida na tela e nos relatórios.
    colunas_numericas = [
        'quantidade_inv', 'quantidade_pos',
        'valor_total_inv', 'valor_total_pos'
    ]
    for coluna in colunas_numericas:
        comparacao[coluna] = pd.to_numeric(comparacao[coluna], errors='coerce').round(2)

    falta_inventario = comparacao['quantidade_inv'].isna()
    falta_posicao = comparacao['quantidade_pos'].isna()
    comparacao['qtd_igual'] = comparacao['quantidade_inv'].eq(comparacao['quantidade_pos'])
    comparacao['valor_igual'] = comparacao['valor_total_inv'].eq(comparacao['valor_total_pos'])

    comparacao['status'] = 'DIFERENCA'
    comparacao.loc[falta_inventario, 'status'] = 'FALTA NO INVENTARIO'
    comparacao.loc[falta_posicao, 'status'] = 'FALTA NA POSICAO'

    itens_presentes = ~falta_inventario & ~falta_posicao
    itens_iguais = itens_presentes & comparacao['qtd_igual'] & comparacao['valor_igual']
    comparacao.loc[itens_iguais, 'status'] = 'OK'

    comparacao['diff_qtd'] = (
        comparacao['quantidade_inv'] - comparacao['quantidade_pos']
    ).round(2)
    comparacao['diff_valor'] = (
        comparacao['valor_total_inv'] - comparacao['valor_total_pos']
    ).round(2)

    return comparacao

def gerar_excel(comparacao):
    relatorio = comparacao[[
        'codigo_formatado_inv',
        'descricao_inv',
        'loja_inv',
        'quantidade_inv',
        'valor_total_inv',
        'quantidade_pos',
        'valor_total_pos',
        'diff_qtd',
        'diff_valor',
        'status'
    ]].copy()

    relatorio.columns = [
        'Codigo',
        'Descricao',
        'Loja',
        'Qtd Inventario',
        'Valor Inventario',
        'Qtd Posicao',
        'Valor Posicao',
        'Diff Qtd',
        'Diff Valor',
        'Status'
    ]

    status_order = {'DIFERENCA': 0, 'FALTA NO INVENTARIO': 1, 'FALTA NA POSICAO': 2, 'OK': 3}
    relatorio['status_order'] = relatorio['Status'].map(status_order)
    relatorio = relatorio.sort_values(['status_order', 'Codigo'])
    relatorio = relatorio.drop('status_order', axis=1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparacao"

    for r_idx, row in enumerate(dataframe_to_rows(relatorio, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif value == 'DIFERENCA':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif value == 'FALTA NO INVENTARIO':
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif value == 'FALTA NA POSICAO':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif value == 'OK':
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if c_idx in [4, 5, 6, 7, 8, 9]:
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal="right")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 18

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

CORES_PDF = {
    'primaria': colors.HexColor('#6366f1'),
    'primaria_escura': colors.HexColor('#4338ca'),
    'texto': colors.HexColor('#1f2937'),
    'texto_claro': colors.HexColor('#6b7280'),
    'linha': colors.HexColor('#e5e7eb'),
    'ok_fundo': colors.HexColor('#E2EFDA'),
    'ok_texto': colors.HexColor('#375623'),
    'diferenca_fundo': colors.HexColor('#FFC7CE'),
    'diferenca_texto': colors.HexColor('#9C0006'),
    'falta_inv_fundo': colors.HexColor('#FFEB9C'),
    'falta_inv_texto': colors.HexColor('#9C6500'),
    'falta_pos_fundo': colors.HexColor('#C6EFCE'),
    'falta_pos_texto': colors.HexColor('#006100'),
}

STATUS_ORDEM_PDF = {'DIFERENCA': 0, 'FALTA NO INVENTARIO': 1, 'FALTA NA POSICAO': 2, 'OK': 3}
STATUS_CORES_PDF = {
    'OK': ('ok_fundo', 'ok_texto'),
    'DIFERENCA': ('diferenca_fundo', 'diferenca_texto'),
    'FALTA NO INVENTARIO': ('falta_inv_fundo', 'falta_inv_texto'),
    'FALTA NA POSICAO': ('falta_pos_fundo', 'falta_pos_texto'),
}

def formatar_numero_pdf(valor, casas=2):
    if valor is None or pd.isna(valor):
        return '-'
    texto = f'{float(valor):,.{casas}f}'
    return texto.replace(',', '§').replace('.', ',').replace('§', '.')

def formatar_moeda_pdf(valor):
    if valor is None or pd.isna(valor):
        return '-'
    return f'R$ {formatar_numero_pdf(valor, 2)}'

class RodapePaginado(Canvas):
    """Canvas que numera as páginas como "Página X de Y" e desenha uma
    faixa de cabeçalho/rodapé de marca em todas as páginas do relatório."""

    def __init__(self, *args, **kwargs):
        Canvas.__init__(self, *args, **kwargs)
        self._paginas_salvas = []

    def showPage(self):
        self._paginas_salvas.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_paginas = len(self._paginas_salvas)
        for estado in self._paginas_salvas:
            self.__dict__.update(estado)
            self._desenhar_moldura(total_paginas)
            Canvas.showPage(self)
        Canvas.save(self)

    def _desenhar_moldura(self, total_paginas):
        largura, altura = self._pagesize
        self.saveState()

        self.setFillColor(CORES_PDF['primaria'])
        self.rect(0, altura - 34, largura, 34, fill=1, stroke=0)
        self.setFillColor(colors.white)
        self.setFont('Helvetica-Bold', 12)
        self.drawString(30, altura - 22, 'Rancho Alegre Produtos Agropecuários')
        self.setFont('Helvetica', 9)
        self.drawRightString(largura - 30, altura - 22, 'Comparador de Inventário')

        self.setStrokeColor(CORES_PDF['linha'])
        self.line(30, 28, largura - 30, 28)
        self.setFillColor(CORES_PDF['texto_claro'])
        self.setFont('Helvetica', 8)
        self.drawString(30, 16, f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        self.drawRightString(largura - 30, 16, f'Página {self._pageNumber} de {total_paginas}')

        self.restoreState()

def cartao_kpi(valor, rotulo, cor_fundo, cor_texto):
    estilo = ParagraphStyle(
        'KPI', fontName='Helvetica', fontSize=8, textColor=cor_texto,
        alignment=TA_CENTER, leading=10
    )
    texto = f'<font size="18" face="Helvetica-Bold">{valor}</font><br/>{rotulo}'
    tabela = Table([[Paragraph(texto, estilo)]], colWidths=[110], rowHeights=[50])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), cor_fundo),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROUNDEDCORNERS', [6, 6, 6, 6]),
    ]))
    return tabela

def grafico_distribuicao_status(contagens):
    """contagens: lista de tuplas (rótulo, valor, cor) com valor > 0."""
    desenho = Drawing(480, 150)

    pizza = Pie()
    pizza.x = 40
    pizza.y = 10
    pizza.width = 130
    pizza.height = 130
    pizza.data = [valor for _, valor, _ in contagens]
    pizza.labels = None
    pizza.simpleLabels = False
    pizza.slices.strokeWidth = 1.2
    pizza.slices.strokeColor = colors.white
    for indice, (_, _, cor) in enumerate(contagens):
        pizza.slices[indice].fillColor = cor
    desenho.add(pizza)

    total = sum(valor for _, valor, _ in contagens)
    legenda = Legend()
    legenda.x = 220
    legenda.y = 115
    legenda.dx = 9
    legenda.dy = 9
    legenda.fontName = 'Helvetica'
    legenda.fontSize = 9
    legenda.leading = 16
    legenda.alignment = 'right'
    legenda.colorNamePairs = [
        (cor, f'{rotulo} — {valor} ({valor / total * 100:.1f}%)')
        for rotulo, valor, cor in contagens
    ]
    desenho.add(legenda)

    return desenho

def gerar_pdf(comparacao, unidade, usuario_nome):
    """Gera o relatório de comparação em PDF: resumo executivo com cartões de
    KPI e gráfico de distribuição, seguido da lista completa de itens que
    precisam de atenção (diferença ou falta). Itens OK só entram na contagem
    do resumo — não há necessidade de listar item a item o que já bate certo.
    """
    output = io.BytesIO()
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(
        output, pagesize=pagesize,
        rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=36
    )
    story = []
    styles = getSampleStyleSheet()

    total = len(comparacao)
    ok_count = len(comparacao[comparacao['status'] == 'OK'])
    diferenca_count = len(comparacao[comparacao['status'] == 'DIFERENCA'])
    falta_inv_count = len(comparacao[comparacao['status'] == 'FALTA NO INVENTARIO'])
    falta_pos_count = len(comparacao[comparacao['status'] == 'FALTA NA POSICAO'])
    taxa = (ok_count / total * 100) if total > 0 else 0

    # Cabeçalho da comparação
    subtitulo_estilo = ParagraphStyle(
        'Subtitulo', fontName='Helvetica', fontSize=10, textColor=CORES_PDF['texto_claro']
    )
    story.append(Paragraph('Relatório de Comparação de Inventário', styles['Heading1']))
    story.append(Paragraph(
        f'Unidade: <b>{unidade}</b> &nbsp;|&nbsp; Gerado por: <b>{usuario_nome}</b> '
        f'&nbsp;|&nbsp; {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        subtitulo_estilo
    ))
    story.append(Spacer(1, 16))

    # Cartões de resumo (KPIs)
    cartoes = [
        cartao_kpi(total, 'Total de Itens', colors.HexColor('#eef2ff'), CORES_PDF['primaria_escura']),
        cartao_kpi(ok_count, 'OK', CORES_PDF['ok_fundo'], CORES_PDF['ok_texto']),
        cartao_kpi(diferenca_count, 'Diferença', CORES_PDF['diferenca_fundo'], CORES_PDF['diferenca_texto']),
        cartao_kpi(falta_inv_count, 'Falta no Inventário', CORES_PDF['falta_inv_fundo'], CORES_PDF['falta_inv_texto']),
        cartao_kpi(falta_pos_count, 'Falta na Posição', CORES_PDF['falta_pos_fundo'], CORES_PDF['falta_pos_texto']),
        cartao_kpi(f'{taxa:.1f}%', 'Taxa de Acurácia', colors.HexColor('#eef2ff'), CORES_PDF['primaria_escura']),
    ]
    linha_cartoes = Table([cartoes], colWidths=[120] * len(cartoes))
    linha_cartoes.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(linha_cartoes)
    story.append(Spacer(1, 16))

    # Gráfico de distribuição (só entram fatias com pelo menos 1 item)
    fatias = [
        ('OK', ok_count, CORES_PDF['ok_texto']),
        ('Diferença', diferenca_count, CORES_PDF['diferenca_texto']),
        ('Falta no Inventário', falta_inv_count, CORES_PDF['falta_inv_texto']),
        ('Falta na Posição', falta_pos_count, CORES_PDF['falta_pos_texto']),
    ]
    fatias = [fatia for fatia in fatias if fatia[1] > 0]
    if fatias:
        story.append(Paragraph('Distribuição por Status', styles['Heading3']))
        story.append(grafico_distribuicao_status(fatias))
        story.append(Spacer(1, 12))

    # Tabela de itens que precisam de atenção (tudo que não é OK)
    pendentes = comparacao[comparacao['status'] != 'OK'].copy()
    pendentes['status_ordem'] = pendentes['status'].map(STATUS_ORDEM_PDF)
    pendentes = pendentes.sort_values(['status_ordem', 'codigo_formatado_inv'])

    story.append(Paragraph(
        f'Itens que Precisam de Atenção ({len(pendentes)})', styles['Heading3']
    ))
    story.append(Spacer(1, 8))

    if len(pendentes) == 0:
        story.append(Paragraph(
            'Nenhuma diferença ou falta encontrada — todos os itens conferem.',
            styles['Normal']
        ))
    else:
        cabecalho = [
            'Código', 'Descrição', 'Loja', 'Qtd Inv', 'Valor Inv',
            'Qtd Pos', 'Valor Pos', 'Diff Qtd', 'Diff Valor', 'Status'
        ]
        dados_tabela = [cabecalho]
        for _, row in pendentes.iterrows():
            dados_tabela.append([
                str(row.get('codigo_formatado_inv', '') or ''),
                str(row.get('descricao_inv', '') or '')[:40],
                formatar_numero_pdf(row.get('loja_inv'), casas=0),
                formatar_numero_pdf(row.get('quantidade_inv')),
                formatar_moeda_pdf(row.get('valor_total_inv')),
                formatar_numero_pdf(row.get('quantidade_pos')),
                formatar_moeda_pdf(row.get('valor_total_pos')),
                formatar_numero_pdf(row.get('diff_qtd')),
                formatar_moeda_pdf(row.get('diff_valor')),
                row['status']
            ])

        larguras = [55, 220, 40, 55, 75, 55, 75, 55, 75, 90]
        tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)

        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), CORES_PDF['primaria']),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (8, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, CORES_PDF['linha']),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]
        for indice, (_, row) in enumerate(pendentes.iterrows(), start=1):
            chave_cor = STATUS_CORES_PDF.get(row['status'])
            if chave_cor:
                fundo, texto = chave_cor
                estilo_tabela.append(('BACKGROUND', (9, indice), (9, indice), CORES_PDF[fundo]))
                estilo_tabela.append(('TEXTCOLOR', (9, indice), (9, indice), CORES_PDF[texto]))
                estilo_tabela.append(('FONTNAME', (9, indice), (9, indice), 'Helvetica-Bold'))

        tabela.setStyle(TableStyle(estilo_tabela))
        story.append(tabela)

    doc.build(story, canvasmaker=RodapePaginado)
    output.seek(0)
    return output

def criar_alerta(comparacao_id, usuario_id, taxa_acuracidade):
    """Cria alertas baseado na taxa de acurácia"""
    limite_critico = obter_configuracao('limiar_alerta_critico', float)
    limite_aviso = obter_configuracao('limiar_alerta_aviso', float)

    if taxa_acuracidade < limite_critico:
        tipo = 'ALERTA_CRITICO'
        limite = limite_critico
    elif taxa_acuracidade < limite_aviso:
        tipo = 'ALERTA_AVISO'
        limite = limite_aviso
    else:
        return

    mensagem = f'Taxa de acurácia abaixo de {limite:g}%: {taxa_acuracidade:.2f}%'
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('''INSERT INTO alertas
                 (comparacao_id, usuario_id, tipo_alerta, mensagem, taxa_acuracidade, limite_alerta)
                 VALUES (?, ?, ?, ?, ?, ?)''',
              (comparacao_id, usuario_id, tipo, mensagem, taxa_acuracidade, limite))

    conn.commit()
    conn.close()

@app.after_request
def desativar_cache(response):
    """Evita que o navegador reutilize uma versão antiga do dashboard/histórico
    (via cache de disco ou back/forward cache) depois de uma nova comparação —
    sintoma: dados corretos no banco, mas a tela mostrando um estado anterior."""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return response

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        senha = request.form.get('senha')

        logger.info(f'Tentativa de login: {username}')

        conn = sqlite3.connect('comparacoes.db')
        c = conn.cursor()
        c.execute('SELECT id, nome, unidade, senha, é_admin FROM usuarios WHERE username = ? AND ativo = 1', (username,))
        usuario = c.fetchone()
        conn.close()

        if usuario and check_password_hash(usuario[3], senha):
            session['usuario_id'] = usuario[0]
            session['nome'] = usuario[1]
            session['unidade'] = usuario[2]
            session['is_admin'] = bool(usuario[4])
            logger.info(f'Login bem-sucedido: {username} (ID: {usuario[0]})')
            return redirect(url_for('dashboard'))
        else:
            logger.warning(f'Falha de login: {username}')
            return render_template('login.html', erro='Usuário ou senha inválidos')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html', unidades=UNIDADES)

@app.route('/api/comparar', methods=['POST'])
@login_required
def api_comparar():
    inv_path = None
    pos_path = None
    try:
        logger.info(f'Iniciando comparação - Usuário: {session.get("usuario_id")}')

        unidade_selecionada = request.form.get('unidade')
        logger.info(f'Unidade selecionada: {unidade_selecionada}')

        if unidade_selecionada not in UNIDADES:
            return jsonify({'erro': 'Unidade inválida'}), 400

        if 'inventario' not in request.files or 'posicao' not in request.files:
            return jsonify({'erro': 'Faltam arquivos'}), 400

        inv_file = request.files['inventario']
        pos_file = request.files['posicao']

        if inv_file.filename == '' or pos_file.filename == '':
            return jsonify({'erro': 'Selecione ambos os arquivos'}), 400

        if not extensao_permitida(inv_file.filename) or not extensao_permitida(pos_file.filename):
            return jsonify({'erro': 'Apenas arquivos .xlsx ou .xls são aceitos'}), 400

        inv_path = os.path.join(UPLOAD_FOLDER, secure_filename('inv_temp.xlsx'))
        pos_path = os.path.join(UPLOAD_FOLDER, secure_filename('pos_temp.xlsx'))

        inv_file.save(inv_path)
        pos_file.save(pos_path)

        logger.info(f'Arquivos salvos: {inv_file.filename}, {pos_file.filename}')

        inv_data = extrair_inventario(inv_path)
        logger.info(f'Inventário extraído: {len(inv_data)} itens')

        lojas_inventario = inv_data['loja'].dropna().unique()
        loja_padrao = lojas_inventario[0] if len(lojas_inventario) == 1 else None
        pos_data = extrair_posicao(pos_path, loja_padrao=loja_padrao)
        logger.info(f'Posição extraída: {len(pos_data)} itens')

        comparacao = comparar(inv_data, pos_data)
        logger.info(f'Comparação concluída')

        ok_count = len(comparacao[comparacao['status'] == 'OK'])
        diff_count = len(comparacao[comparacao['status'] == 'DIFERENCA'])
        falta_inv = len(comparacao[comparacao['status'] == 'FALTA NO INVENTARIO'])
        falta_pos = len(comparacao[comparacao['status'] == 'FALTA NA POSICAO'])

        relatorio_display = comparacao[[
            'codigo_formatado_inv',
            'descricao_inv',
            'loja_inv',
            'quantidade_inv',
            'valor_total_inv',
            'quantidade_pos',
            'valor_total_pos',
            'diff_qtd',
            'diff_valor',
            'status'
        ]].copy()

        relatorio_display.columns = [
            'Codigo', 'Descricao', 'Loja', 'Qtd Inv', 'Valor Inv',
            'Qtd Pos', 'Valor Pos', 'Diff Qtd', 'Diff Valor', 'Status'
        ]

        # O merge sempre deixa os itens que só existem na posição (FALTA NO
        # INVENTARIO) no final da lista. Se a comparação passar do limite de
        # exibição, eles nunca apareceriam na tabela do site. Ordenar por
        # prioridade de status antes de cortar garante que diferenças/faltas
        # sempre apareçam primeiro — só os itens "OK" ficam de fora do corte.
        relatorio_display['status_order'] = relatorio_display['Status'].map(STATUS_ORDEM_PDF)
        relatorio_display = relatorio_display.sort_values(['status_order', 'Codigo'])
        relatorio_display = relatorio_display.drop(columns='status_order')

        relatorio_display = relatorio_display.fillna('')
        relatorio_json = relatorio_display.to_json(orient='records')

        # Salvar no banco
        conn = sqlite3.connect('comparacoes.db')
        c = conn.cursor()
        c.execute('''INSERT INTO comparacoes
                     (usuario_id, unidade, arquivo_inventario, arquivo_posicao, total_itens, itens_iguais, itens_diferentes, itens_falta, resultado_json)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (session['usuario_id'], unidade_selecionada, inv_file.filename, pos_file.filename,
                   len(comparacao), ok_count, diff_count, falta_inv + falta_pos, relatorio_json))

        comparacao_id = c.lastrowid

        # Salvar detalhes
        for _, row in comparacao.iterrows():
            c.execute('''INSERT INTO comparacao_detalhes
                         (comparacao_id, codigo, descricao, loja, qtd_inv, valor_inv, qtd_pos, valor_pos, diff_qtd, diff_valor, status)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                      (comparacao_id, row.get('codigo_formatado_inv', ''), row.get('descricao_inv', ''),
                       row.get('loja_inv', ''), row.get('quantidade_inv'), row.get('valor_total_inv'),
                       row.get('quantidade_pos'), row.get('valor_total_pos'), row.get('diff_qtd'), row.get('diff_valor'),
                       row['status']))

        conn.commit()

        # Criar alertas se necessário
        taxa_acuracidade = round((ok_count / len(comparacao) * 100), 2) if len(comparacao) > 0 else 0
        criar_alerta(comparacao_id, session['usuario_id'], taxa_acuracidade)

        conn.close()

        session['comparacao_id'] = comparacao_id

        resultado = {
            'id': comparacao_id,
            'total': len(comparacao),
            'iguais': ok_count,
            'diferentes': diff_count,
            'falta_inv': falta_inv,
            'falta_pos': falta_pos,
            'taxa_acuracidade': taxa_acuracidade,
            'dados': json.loads(relatorio_json)[:obter_configuracao('limite_resultados', int)]
        }

        logger.info(f'Comparação concluída com sucesso: {len(comparacao)} itens')
        return jsonify(resultado)

    except ValueError as e:
        logger.warning(f'Arquivo incompatível na comparação: {str(e)}')
        return jsonify({'erro': str(e)}), 400
    except Exception as e:
        logger.error(f'Erro na comparação: {str(e)}', exc_info=True)
        return jsonify({'erro': str(e)}), 500
    finally:
        for caminho_temporario in (inv_path, pos_path):
            if caminho_temporario and os.path.exists(caminho_temporario):
                try:
                    os.remove(caminho_temporario)
                except OSError:
                    logger.warning(f'Não foi possível remover o arquivo temporário: {caminho_temporario}')

@app.route('/api/download-excel', defaults={'id': None}, methods=['GET'])
@app.route('/api/download-excel/<int:id>', methods=['GET'])
@login_required
def download_excel(id):
    comparacao_id = id or session.get('comparacao_id')
    if not comparacao_id:
        return jsonify({'erro': 'Nenhuma comparação selecionada'}), 400

    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        'SELECT unidade FROM comparacoes WHERE id = ? AND usuario_id = ?',
        (comparacao_id, session['usuario_id'])
    )
    comp = c.fetchone()
    if not comp:
        conn.close()
        return jsonify({'erro': 'Comparação não encontrada'}), 404

    c.execute(
        '''SELECT codigo, descricao, loja, qtd_inv, valor_inv, qtd_pos,
                  valor_pos, diff_qtd, diff_valor, status
           FROM comparacao_detalhes WHERE comparacao_id = ?''',
        (comparacao_id,)
    )
    detalhes = c.fetchall()
    conn.close()

    dados = [{
        'codigo_formatado_inv': row['codigo'],
        'descricao_inv': row['descricao'],
        'loja_inv': row['loja'],
        'quantidade_inv': row['qtd_inv'],
        'valor_total_inv': row['valor_inv'],
        'quantidade_pos': row['qtd_pos'],
        'valor_total_pos': row['valor_pos'],
        'diff_qtd': row['diff_qtd'],
        'diff_valor': row['diff_valor'],
        'status': row['status']
    } for row in detalhes]

    colunas = [
        'codigo_formatado_inv', 'descricao_inv', 'loja_inv',
        'quantidade_inv', 'valor_total_inv', 'quantidade_pos',
        'valor_total_pos', 'diff_qtd', 'diff_valor', 'status'
    ]
    excel_output = gerar_excel(pd.DataFrame(dados, columns=colunas))
    return send_file(
        excel_output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Comparacao_{comp['unidade']}_{datetime.now().strftime('%d_%m_%Y_%H%M%S')}.xlsx"
    )

@app.route('/api/historico', methods=['GET'])
@login_required
def get_historico():
    unidade_filtro = request.args.get('unidade', '')
    limite_historico = obter_configuracao('limite_historico', int)

    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    if unidade_filtro and unidade_filtro in UNIDADES:
        c.execute('''SELECT id, data_criacao, unidade, arquivo_inventario, arquivo_posicao, total_itens, itens_iguais, itens_diferentes
                     FROM comparacoes WHERE usuario_id = ? AND unidade = ? ORDER BY id DESC LIMIT ?''',
                  (session['usuario_id'], unidade_filtro, limite_historico))
    else:
        c.execute('''SELECT id, data_criacao, unidade, arquivo_inventario, arquivo_posicao, total_itens, itens_iguais, itens_diferentes
                     FROM comparacoes WHERE usuario_id = ? ORDER BY id DESC LIMIT ?''',
                  (session['usuario_id'], limite_historico))

    rows = c.fetchall()
    conn.close()

    historico = []
    for row in rows:
        historico.append({
            'id': row['id'],
            'data': row['data_criacao'],
            'unidade': row['unidade'],
            'inv': row['arquivo_inventario'],
            'pos': row['arquivo_posicao'],
            'total': row['total_itens'],
            'iguais': row['itens_iguais'],
            'diferentes': row['itens_diferentes']
        })

    return jsonify(historico)

@app.route('/api/comparacao/<int:id>', methods=['GET'])
@login_required
def get_comparacao(id):
    limite_resultados = obter_configuracao('limite_resultados', int)
    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute('SELECT * FROM comparacoes WHERE id = ? AND usuario_id = ?', (id, session['usuario_id']))
    comp = c.fetchone()

    if not comp:
        conn.close()
        return jsonify({'erro': 'Comparação não encontrada'}), 404

    c.execute('SELECT * FROM comparacao_detalhes WHERE comparacao_id = ? LIMIT ?', (id, limite_resultados))
    detalhes = c.fetchall()

    conn.close()

    dados = [{
        'Codigo': row['codigo'],
        'Descricao': row['descricao'],
        'Loja': row['loja'],
        'Qtd Inv': row['qtd_inv'],
        'Valor Inv': row['valor_inv'],
        'Qtd Pos': row['qtd_pos'],
        'Valor Pos': row['valor_pos'],
        'Diff Qtd': row['diff_qtd'],
        'Diff Valor': row['diff_valor'],
        'Status': row['status']
    } for row in detalhes]

    return jsonify({
        'id': comp['id'],
        'data': comp['data_criacao'],
        'unidade': comp['unidade'],
        'total': comp['total_itens'],
        'iguais': comp['itens_iguais'],
        'diferentes': comp['itens_diferentes'],
        'falta': comp['itens_falta'],
        'dados': dados
    })

@app.route('/api/estatisticas', methods=['GET'])
@login_required
def get_estatisticas():
    unidade_filtro = request.args.get('unidade', '')

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()

    if unidade_filtro and unidade_filtro in UNIDADES:
        c.execute('''SELECT COUNT(*) as total, SUM(itens_iguais) as iguais, SUM(itens_diferentes) as diferentes
                     FROM comparacoes WHERE usuario_id = ? AND unidade = ? AND data_criacao >= datetime('now', '-30 days')''',
                  (session['usuario_id'], unidade_filtro))
    else:
        c.execute('''SELECT COUNT(*) as total, SUM(itens_iguais) as iguais, SUM(itens_diferentes) as diferentes
                     FROM comparacoes WHERE usuario_id = ? AND data_criacao >= datetime('now', '-30 days')''',
                  (session['usuario_id'],))

    result = c.fetchone()
    conn.close()

    total_comp = result[0] or 0
    total_iguais = result[1] or 0
    total_diferentes = result[2] or 0

    return jsonify({
        'total_comparacoes': total_comp,
        'total_itens_iguais': total_iguais,
        'total_itens_diferentes': total_diferentes,
        'taxa_media': round((total_iguais / (total_iguais + total_diferentes) * 100), 2) if (total_iguais + total_diferentes) > 0 else 0
    })

def validar_senha(senha):
    """Validar força da senha no servidor"""
    if len(senha) < 8:
        return False, 'Senha deve ter pelo menos 8 caracteres'
    if not any(c.isupper() for c in senha):
        return False, 'Senha deve conter uma letra maiúscula'
    if not any(c.isdigit() for c in senha):
        return False, 'Senha deve conter um número'
    if not any(c in '!@#$%^&*(),.?":{}|<>' for c in senha):
        return False, 'Senha deve conter um caractere especial (!@#$%)'
    return True, 'OK'

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username')
        nome = request.form.get('nome')
        email = request.form.get('email')
        senha = request.form.get('senha')
        unidade = request.form.get('unidade')

        if not all([username, nome, senha, unidade]):
            return render_template('signup.html', erro='Preencha todos os campos', unidades=UNIDADES)

        if unidade not in UNIDADES:
            return render_template('signup.html', erro='Unidade inválida', unidades=UNIDADES)

        # Validar força da senha no servidor
        valido, mensagem = validar_senha(senha)
        if not valido:
            return render_template('signup.html', erro=mensagem, unidades=UNIDADES)

        conn = sqlite3.connect('comparacoes.db')
        c = conn.cursor()

        try:
            senha_hash = generate_password_hash(senha)
            c.execute('''INSERT INTO usuarios (username, senha, nome, email, unidade)
                         VALUES (?, ?, ?, ?, ?)''',
                      (username, senha_hash, nome, email, unidade))
            conn.commit()
            conn.close()

            return render_template('signup.html', sucesso='Usuário criado com sucesso! Faça login agora.', unidades=UNIDADES)

        except sqlite3.IntegrityError:
            conn.close()
            return render_template('signup.html', erro='Usuário já existe', unidades=UNIDADES)
        except Exception as e:
            conn.close()
            return render_template('signup.html', erro=str(e), unidades=UNIDADES)

    return render_template('signup.html', unidades=UNIDADES)

@app.route('/admin')
@login_required
def admin():
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()

    # Verificar se é admin
    c.execute('SELECT é_admin, nome FROM usuarios WHERE id = ?', (session['usuario_id'],))
    user = c.fetchone()

    if not user or not user[0]:
        conn.close()
        return redirect(url_for('dashboard'))

    # Contar estatísticas
    c.execute('SELECT COUNT(*) FROM comparacoes')
    total_comparacoes = c.fetchone()[0]

    c.execute('SELECT COUNT(*) FROM alertas')
    total_alertas = c.fetchone()[0]

    conn.close()

    return render_template('admin.html',
                         usuario=user[1],
                         total_comparacoes=total_comparacoes,
                         total_alertas=total_alertas,
                         unidades=UNIDADES)

@app.route('/api/usuarios', methods=['GET'])
@login_required
def api_usuarios():
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute(
        '''SELECT id, username, nome, email, unidade, ativo, é_admin, criado_em
           FROM usuarios ORDER BY id DESC'''
    )
    usuarios = []
    for row in c.fetchall():
        usuarios.append({
            'id': row[0],
            'username': row[1],
            'nome': row[2],
            'email': row[3],
            'unidade': row[4],
            'ativo': bool(row[5]),
            'admin': bool(row[6]),
            'criado_em': row[7]
        })

    conn.close()
    return jsonify(usuarios)

@app.route('/api/admin/adicionar-usuario', methods=['POST'])
@login_required
def adicionar_usuario_admin():
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()

    username = (request.form.get('username') or '').strip()
    nome = (request.form.get('nome') or '').strip()
    email = (request.form.get('email') or '').strip()
    unidade = request.form.get('unidade') or ''
    senha = request.form.get('senha') or ''

    if not all([username, nome, unidade, senha]):
        conn.close()
        return jsonify({'erro': 'Preencha todos os campos obrigatórios'}), 400

    if unidade not in UNIDADES:
        conn.close()
        return jsonify({'erro': 'Unidade inválida'}), 400

    valido, mensagem = validar_senha(senha)
    if not valido:
        conn.close()
        return jsonify({'erro': mensagem}), 400

    try:
        c.execute(
            '''INSERT INTO usuarios (username, senha, nome, email, unidade)
               VALUES (?, ?, ?, ?, ?)''',
            (username, generate_password_hash(senha), nome, email, unidade)
        )
        novo_usuario_id = c.lastrowid
        conn.commit()
        registrar_atividade(
            session['usuario_id'],
            'USUARIO_CRIADO',
            f'Usuário {username} criado (ID {novo_usuario_id})'
        )
        return jsonify({'sucesso': True, 'mensagem': 'Usuário criado com sucesso'})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Usuário já existe'}), 409
    finally:
        conn.close()

@app.route('/api/usuario/<int:id>', methods=['PUT'])
@login_required
def editar_usuario(id):
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    dados = request.get_json(silent=True) or request.form
    username = (dados.get('username') or '').strip()
    nome = (dados.get('nome') or '').strip()
    email = (dados.get('email') or '').strip()
    unidade = dados.get('unidade') or ''
    senha = dados.get('senha') or ''
    admin = str(dados.get('admin', '')).lower() in ('1', 'true', 'on', 'sim')

    if not all([username, nome, unidade]):
        return jsonify({'erro': 'Preencha todos os campos obrigatórios'}), 400
    if unidade not in UNIDADES:
        return jsonify({'erro': 'Unidade inválida'}), 400
    if senha:
        valido, mensagem = validar_senha(senha)
        if not valido:
            return jsonify({'erro': mensagem}), 400
    if id == session['usuario_id'] and not admin:
        return jsonify({'erro': 'Você não pode remover seu próprio acesso administrativo'}), 400

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('SELECT username FROM usuarios WHERE id = ?', (id,))
    usuario_atual = c.fetchone()
    if not usuario_atual:
        conn.close()
        return jsonify({'erro': 'Usuário não encontrado'}), 404

    try:
        parametros = [username, nome, email, unidade, int(admin)]
        query = '''UPDATE usuarios
                   SET username = ?, nome = ?, email = ?, unidade = ?, é_admin = ?'''
        if senha:
            query += ', senha = ?'
            parametros.append(generate_password_hash(senha))
        query += ' WHERE id = ?'
        parametros.append(id)

        c.execute(query, parametros)
        conn.commit()

        if id == session['usuario_id']:
            session['nome'] = nome
            session['unidade'] = unidade
            session['is_admin'] = admin

        registrar_atividade(
            session['usuario_id'],
            'USUARIO_EDITADO',
            f'Usuário {usuario_atual[0]} atualizado para {username}'
        )
        return jsonify({'sucesso': True, 'mensagem': 'Usuário atualizado com sucesso'})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Este nome de usuário já está em uso'}), 409
    finally:
        conn.close()

@app.route('/api/usuario/<int:id>/toggle', methods=['POST'])
@login_required
def toggle_usuario(id):
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403
    if id == session['usuario_id']:
        return jsonify({'erro': 'Você não pode desativar seu próprio usuário'}), 400

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('SELECT username, ativo FROM usuarios WHERE id = ?', (id,))
    usuario = c.fetchone()
    if not usuario:
        conn.close()
        return jsonify({'erro': 'Usuário não encontrado'}), 404

    novo_status = not bool(usuario[1])
    c.execute('UPDATE usuarios SET ativo = ? WHERE id = ?', (int(novo_status), id))
    conn.commit()
    conn.close()

    registrar_atividade(
        session['usuario_id'],
        'USUARIO_ATIVADO' if novo_status else 'USUARIO_DESATIVADO',
        f'Usuário {usuario[0]} foi {"ativado" if novo_status else "desativado"}'
    )
    return jsonify({'sucesso': True, 'ativo': novo_status})

@app.route('/api/usuario/<int:id>', methods=['DELETE'])
@login_required
def excluir_usuario(id):
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403
    if id == session['usuario_id']:
        return jsonify({'erro': 'Você não pode excluir seu próprio usuário'}), 400

    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    c.execute('SELECT username, é_admin, ativo FROM usuarios WHERE id = ?', (id,))
    usuario = c.fetchone()
    if not usuario:
        conn.close()
        return jsonify({'erro': 'Usuário não encontrado'}), 404

    if usuario[1] and usuario[2]:
        c.execute('SELECT COUNT(*) FROM usuarios WHERE é_admin = 1 AND ativo = 1 AND id != ?', (id,))
        outros_admins_ativos = c.fetchone()[0]
        if outros_admins_ativos == 0:
            conn.close()
            return jsonify({'erro': 'Não é possível excluir o único administrador ativo'}), 400

    c.execute('DELETE FROM usuarios WHERE id = ?', (id,))
    conn.commit()
    conn.close()

    registrar_atividade(
        session['usuario_id'],
        'USUARIO_EXCLUIDO',
        f'Usuário {usuario[0]} foi excluído permanentemente'
    )
    return jsonify({'sucesso': True})

@app.route('/api/admin/estatisticas', methods=['GET'])
@login_required
def estatisticas_admin():
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute(
        '''SELECT COUNT(*) total,
                  SUM(CASE WHEN ativo = 1 THEN 1 ELSE 0 END) ativos,
                  SUM(CASE WHEN criado_em >= datetime('now', '-7 days') THEN 1 ELSE 0 END) novos_semana
           FROM usuarios'''
    )
    usuarios = c.fetchone()

    c.execute(
        '''SELECT COUNT(*) total_comparacoes,
                  COALESCE(SUM(total_itens), 0) total_itens,
                  COALESCE(SUM(itens_iguais), 0) itens_iguais,
                  COALESCE(SUM(itens_diferentes + itens_falta), 0) divergencias,
                  SUM(CASE WHEN data_criacao >= datetime('now', '-30 days') THEN 1 ELSE 0 END) comparacoes_30_dias
           FROM comparacoes'''
    )
    comparacoes = c.fetchone()

    c.execute('SELECT COUNT(*) total, SUM(CASE WHEN lido = 0 THEN 1 ELSE 0 END) nao_lidos FROM alertas')
    alertas = c.fetchone()

    c.execute(
        '''SELECT unidade,
                  COUNT(*) comparacoes,
                  COALESCE(SUM(total_itens), 0) total_itens,
                  COALESCE(SUM(itens_iguais), 0) itens_iguais,
                  COALESCE(SUM(itens_diferentes + itens_falta), 0) divergencias
           FROM comparacoes
           GROUP BY unidade
           ORDER BY unidade'''
    )
    por_unidade = []
    for row in c.fetchall():
        total_itens = row['total_itens'] or 0
        por_unidade.append({
            'unidade': row['unidade'],
            'comparacoes': row['comparacoes'],
            'total_itens': total_itens,
            'itens_iguais': row['itens_iguais'],
            'divergencias': row['divergencias'],
            'acuracidade': round(row['itens_iguais'] / total_itens * 100, 2) if total_itens else 0
        })

    c.execute(
        '''SELECT date(data_criacao) data,
                  COUNT(*) comparacoes,
                  COALESCE(SUM(total_itens), 0) itens
           FROM comparacoes
           WHERE data_criacao >= datetime('now', '-13 days')
           GROUP BY date(data_criacao)
           ORDER BY data'''
    )
    por_dia = [dict(row) for row in c.fetchall()]
    conn.close()

    total_itens = comparacoes['total_itens'] or 0
    return jsonify({
        'resumo': {
            'usuarios': usuarios['total'] or 0,
            'usuarios_ativos': usuarios['ativos'] or 0,
            'usuarios_novos_semana': usuarios['novos_semana'] or 0,
            'comparacoes': comparacoes['total_comparacoes'] or 0,
            'comparacoes_30_dias': comparacoes['comparacoes_30_dias'] or 0,
            'total_itens': total_itens,
            'itens_iguais': comparacoes['itens_iguais'] or 0,
            'divergencias': comparacoes['divergencias'] or 0,
            'acuracidade': round(comparacoes['itens_iguais'] / total_itens * 100, 2) if total_itens else 0,
            'alertas': alertas['total'] or 0,
            'alertas_nao_lidos': alertas['nao_lidos'] or 0
        },
        'por_unidade': por_unidade,
        'por_dia': por_dia
    })

@app.route('/api/admin/atividades', methods=['GET'])
@login_required
def atividades_admin():
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        '''SELECT data, tipo, usuario, unidade, mensagem FROM (
               SELECT a.criado_em data, a.tipo tipo,
                      COALESCE(u.nome, 'Sistema') usuario,
                      COALESCE(u.unidade, '-') unidade,
                      a.mensagem mensagem
               FROM atividades a
               LEFT JOIN usuarios u ON u.id = a.usuario_id

               UNION ALL

               SELECT c.data_criacao data, 'COMPARACAO' tipo,
                      COALESCE(u.nome, 'Usuário removido') usuario,
                      c.unidade unidade,
                      'Comparação #' || c.id || ' processada: ' || c.total_itens || ' itens' mensagem
               FROM comparacoes c
               LEFT JOIN usuarios u ON u.id = c.usuario_id

               UNION ALL

               SELECT al.criado_em data, al.tipo_alerta tipo,
                      COALESCE(u.nome, 'Usuário removido') usuario,
                      COALESCE(u.unidade, '-') unidade,
                      al.mensagem mensagem
               FROM alertas al
               LEFT JOIN usuarios u ON u.id = al.usuario_id
           )
           ORDER BY data DESC
           LIMIT 100'''
    )
    atividades = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(atividades)

@app.route('/api/admin/configuracoes', methods=['GET', 'POST'])
@login_required
def configuracoes_admin():
    if not verificar_admin():
        return jsonify({'erro': 'Acesso negado'}), 403

    if request.method == 'GET':
        return jsonify({
            'limiar_alerta_critico': obter_configuracao('limiar_alerta_critico', float),
            'limiar_alerta_aviso': obter_configuracao('limiar_alerta_aviso', float),
            'limite_historico': obter_configuracao('limite_historico', int),
            'limite_resultados': obter_configuracao('limite_resultados', int)
        })

    dados = request.get_json(silent=True) or request.form
    try:
        critico = float(dados.get('limiar_alerta_critico'))
        aviso = float(dados.get('limiar_alerta_aviso'))
        limite_historico = int(dados.get('limite_historico'))
        limite_resultados = int(dados.get('limite_resultados'))
    except (TypeError, ValueError):
        return jsonify({'erro': 'Preencha as configurações com números válidos'}), 400

    if not 0 <= critico < aviso <= 100:
        return jsonify({'erro': 'O limite crítico deve ser menor que o limite de aviso'}), 400
    if not 5 <= limite_historico <= 100:
        return jsonify({'erro': 'O histórico deve ficar entre 5 e 100 registros'}), 400
    if not 100 <= limite_resultados <= 5000:
        return jsonify({'erro': 'O limite de resultados deve ficar entre 100 e 5000 itens'}), 400

    valores = {
        'limiar_alerta_critico': str(critico),
        'limiar_alerta_aviso': str(aviso),
        'limite_historico': str(limite_historico),
        'limite_resultados': str(limite_resultados)
    }
    conn = sqlite3.connect('comparacoes.db')
    c = conn.cursor()
    for chave, valor in valores.items():
        c.execute(
            '''INSERT INTO configuracoes (chave, valor, atualizado_em)
               VALUES (?, ?, CURRENT_TIMESTAMP)
               ON CONFLICT(chave) DO UPDATE SET
                   valor = excluded.valor,
                   atualizado_em = CURRENT_TIMESTAMP''',
            (chave, valor)
        )
    conn.commit()
    conn.close()

    registrar_atividade(
        session['usuario_id'],
        'CONFIGURACOES_ATUALIZADAS',
        'Configurações administrativas atualizadas'
    )
    return jsonify({'sucesso': True, 'mensagem': 'Configurações salvas com sucesso'})

@app.route('/api/download-pdf/<int:id>', methods=['GET'])
@login_required
def download_pdf(id):
    try:
        conn = sqlite3.connect('comparacoes.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        c.execute('SELECT * FROM comparacoes WHERE id = ? AND usuario_id = ?', (id, session['usuario_id']))
        comp = c.fetchone()

        if not comp:
            conn.close()
            return jsonify({'erro': 'Comparação não encontrada'}), 404

        c.execute('SELECT * FROM comparacao_detalhes WHERE comparacao_id = ?', (id,))
        detalhes = c.fetchall()

        conn.close()

        # Construir DataFrame com os dados
        dados = []
        for row in detalhes:
            dados.append({
                'codigo_formatado_inv': row['codigo'],
                'descricao_inv': row['descricao'],
                'loja_inv': row['loja'],
                'quantidade_inv': row['qtd_inv'],
                'valor_total_inv': row['valor_inv'],
                'quantidade_pos': row['qtd_pos'],
                'valor_total_pos': row['valor_pos'],
                'diff_qtd': row['diff_qtd'],
                'diff_valor': row['diff_valor'],
                'status': row['status']
            })

        df = pd.DataFrame(dados)

        pdf_output = gerar_pdf(df, comp['unidade'], session['nome'])

        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Relatorio_{comp['unidade']}_{datetime.now().strftime('%d_%m_%Y')}.pdf"
        )

    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/api/alertas', methods=['GET'])
@login_required
def get_alertas():
    conn = sqlite3.connect('comparacoes.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute('''SELECT id, tipo_alerta, mensagem, taxa_acuracidade, limite_alerta, criado_em, lido
                 FROM alertas WHERE usuario_id = ? ORDER BY criado_em DESC LIMIT 10''',
              (session['usuario_id'],))

    alertas = []
    for row in c.fetchall():
        alertas.append({
            'id': row['id'],
            'tipo': row['tipo_alerta'],
            'mensagem': row['mensagem'],
            'taxa': f"{row['taxa_acuracidade']:.2f}%",
            'limite': row['limite_alerta'],
            'data': row['criado_em'],
            'lido': bool(row['lido'])
        })

    conn.close()
    return jsonify(alertas)

@app.route('/api/comparacoes-filtrado', methods=['GET'])
@login_required
def get_comparacoes_filtrado():
    try:
        status_filtro = request.args.get('status', '')
        descricao_filtro = request.args.get('descricao', '').strip()
        comparacao_id = request.args.get('comparacao_id', '')

        if not comparacao_id:
            return jsonify({'erro': 'ID da comparação necessário'}), 400

        conn = sqlite3.connect('comparacoes.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()

        # Verificar permissão
        c.execute('SELECT usuario_id FROM comparacoes WHERE id = ?', (comparacao_id,))
        comp = c.fetchone()

        if not comp or comp['usuario_id'] != session['usuario_id']:
            conn.close()
            return jsonify({'erro': 'Acesso negado'}), 403

        # Construir query com filtros
        query = 'SELECT * FROM comparacao_detalhes WHERE comparacao_id = ?'
        params = [comparacao_id]

        if status_filtro:
            query += ' AND status = ?'
            params.append(status_filtro)

        if descricao_filtro:
            query += ' AND descricao LIKE ? COLLATE NOCASE'
            params.append(f'%{descricao_filtro}%')

        query += ' LIMIT ?'
        params.append(obter_configuracao('limite_resultados', int))

        c.execute(query, params)
        detalhes = c.fetchall()
        conn.close()

        dados = []
        for row in detalhes:
            dados.append({
                'Codigo': row['codigo'],
                'Descricao': row['descricao'],
                'Loja': row['loja'],
                'Qtd Inv': row['qtd_inv'],
                'Valor Inv': row['valor_inv'],
                'Qtd Pos': row['qtd_pos'],
                'Valor Pos': row['valor_pos'],
                'Diff Qtd': row['diff_qtd'],
                'Diff Valor': row['diff_valor'],
                'Status': row['status']
            })

        return jsonify(dados)

    except Exception as e:
        return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    host = os.environ.get('HOST', 'localhost')
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_mode, host=host, port=port)
